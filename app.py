import flask
from flask import Flask, jsonify, request
import json,os,shutil
from src.bin.vf_blp import *
from pandas import ExcelWriter
import pickle
import base64
app = Flask(__name__)
total_yrs = 0
@app.route("/", methods=["GET","POST"])
def run_ml():
    global total_yrs
    if flask.request.method == 'GET':
        cwd=os.getcwd()#current directory
        os.chdir(cwd)
        #print(cwd)
        
        path=cwd+'/'+ 'greenfield'
        prep_input(flask.request,path)
        # get all argument and parse them to uc6_run()
        output_excel = run_uc6(os.path.join(path,"greenfield.json"))
        file_name = os.path.basename(output_excel)
        
        # convert result to json
        multi_sheet_file = pd.ExcelFile(output_excel )
        excel_sheet_names = multi_sheet_file.sheet_names
        dict = {}
        for sheet in excel_sheet_names:
            df = pd.read_excel(os.path.join(path,output_excel),index_col=False,keep_default_na=True)
            dict[sheet] = df
        # Use dumps() to make it serialized
        pickled = pickle.dumps(dict)
        pickled_b64 = base64.b64encode(pickled)
        pickled_b64_str = pickled_b64.decode('utf-8')
        
        # data = json.loads()
        # return json
    return jsonify({file_name:pickled_b64_str})
'" prepare the input file in excel format as the'
'" ML script takes the input as excel file only"'
def getExcelCol(n,let):
    d = n-26
    n= n-1
    rem = int(n%26)+1
    dev = int(n/26)
    str = ''
    str1= ''
    if dev>0:
        str = chr(ord('@')+dev)
    if rem >0:
        str1 = (chr(ord('A') + rem-1 ))
    return(str+str1)
def prep_input(request,path):
    if  os.path.isdir(path):
        shutil.rmtree(path)
    os.mkdir('greenfield')
    #create greenfiled.json file
    data = request.get_json()
    res = json.loads(data)
    print(type(res))
    res["base"]= path
    with open(os.path.join(path,"greenfield.json"), "w+") as f:
        json.dump(res, f)    
    
    '''create vbom excel file from the json data '''
    pickled_b64_str = res["excel_data"]
    byte_pickled = pickled_b64_str.encode('utf-8')
    dict = pickle.loads(base64.b64decode(byte_pickled))#base64.b64decode(hug_pickled.encode())
    writer = ExcelWriter(os.path.join(path, res["excel"]), engine='openpyxl')

    for key,val in dict.items():
        total_yrs = int((len(val.columns)-9)/17)
        
        val.to_excel(writer,index=False,sheet_name=key)
    writer.save()
    #open the workbook and merge the cells.
    import openpyxl
    workbook = openpyxl.load_workbook(os.path.join(path, res["excel"]))
    for sheet in workbook.sheetnames:
        sheet = workbook[sheet]          
            
        n = 9
        k = 9
        for i in range(total_yrs):
            k= k+17
            let = 'I'
            print(getExcelCol(n+1,let)+'2'+':'+getExcelCol(k,let)+'2')
            sheet.merge_cells('A2:I2')
            sheet.merge_cells(getExcelCol(n+1,let)+'2'+':'+getExcelCol(k,let)+'2')
            n += 17
    workbook.save(os.path.join(path, res["excel"]))
        # sheet.merge_cells('J2:Z2')
        # sheet.merge_cells('AA2:AQ2')
        # sheet.merge_cells('AR2:BH2')
        # sheet.merge_cells('BI2:BY2')
        # sheet.merge_cells('BZ2:CP2')
        # sheet.merge_cells('CQ2:DG2')
        # sheet.merge_cells('DH2:DX2')
    '''create CATALOG excel file'''
    pickled_b64_str = res["catalog_data"]
    byte_pickled = pickled_b64_str.encode('utf-8')
    dict = pickle.loads(base64.b64decode(byte_pickled))#base64.b64decode(hug_pickled.encode())
    writer = ExcelWriter(os.path.join(path, res["catalog"]), engine='openpyxl')
    for key,val in dict.items():
        val.to_excel(writer,index=False,sheet_name=key)
    writer.save()

    '''create Cluster Config json file '''
    result = json.loads(res['config_data'])
    json_object = json.dumps(result, indent = 4)
    with open(os.path.join(path, res["config"]), "w") as outfile:
        outfile.write(json_object)

    file_path =''
    return file_path
if __name__ == '__main__':
    app.run(host='127.0.0.1',port=5000,debug=True)